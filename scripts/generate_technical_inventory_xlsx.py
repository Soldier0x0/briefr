#!/usr/bin/env python3
"""Generate TECHNICAL_INVENTORY.xlsx from structured data."""

from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("Install openpyxl: pip install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "TECHNICAL_INVENTORY.xlsx"


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min(max_len + 2, 60), 10)


def main():
    wb = Workbook()

    # Sheet 1: Tech Stack
    ws = wb.active
    ws.title = "Tech Stack"
    ws.append(["Component", "Technology", "Version", "Purpose"])
    for row in [
        ("API framework", "FastAPI", "0.136.3", "REST API, OpenAPI, validation"),
        ("ASGI server", "uvicorn", "0.48.0", "Production HTTP server"),
        ("HTTP client", "httpx", "0.28.1", "Async external API calls"),
        ("Scheduler", "APScheduler", "3.11.2", "7 background jobs"),
        ("Database", "SQLite + aiosqlite", "0.22.1", "Local persistence"),
        ("Validation", "Pydantic", "2.13.4", "Request models"),
        ("UI", "React", "18.3.1", "Analyst SPA"),
        ("Build", "Vite", "5.4.1", "Dev and production bundle"),
        ("PDF", "jsPDF + html2canvas", "4.2.1 / 1.4.1", "Client PDF reports"),
        ("Spreadsheet generation", "openpyxl", "3.1.5", "TECHNICAL_INVENTORY.xlsx script"),
    ]:
        ws.append(row)
    autosize(ws)

    # Sheet 2: Database Schema
    ws2 = wb.create_sheet("Database Schema")
    ws2.append(["Table", "Column", "Type", "Constraints", "Description"])
    schema = [
        # cves (18 columns)
        ("cves", "cve_id", "TEXT", "PRIMARY KEY", "CVE identifier"),
        ("cves", "description", "TEXT", "", "NVD English description"),
        ("cves", "cvss_score", "REAL", "", "CVSS v3 base score"),
        ("cves", "severity", "TEXT", "", "CRITICAL/HIGH/MEDIUM/LOW"),
        ("cves", "published", "TEXT", "", "ISO publish timestamp"),
        ("cves", "modified", "TEXT", "", "ISO last-modified"),
        ("cves", "affected_products", "TEXT", "DEFAULT '[]'", "JSON array vendor:product"),
        ("cves", "mitre_technique", "TEXT", "", "Primary ATT&CK ID from refs"),
        ("cves", "summary", "TEXT", "", "Plain-English summary (KEV/OSV)"),
        ("cves", "is_kev", "INTEGER", "DEFAULT 0", "CISA KEV flag"),
        ("cves", "epss_score", "REAL", "", "Latest EPSS probability"),
        ("cves", "has_poc", "INTEGER", "DEFAULT 0", "Public PoC/exploit flag"),
        ("cves", "patch_available", "INTEGER", "DEFAULT 0", "Patch reference detected"),
        ("cves", "has_ai_context", "INTEGER", "DEFAULT 0", "AI/ML relevance flag"),
        ("cves", "source_urls", "TEXT", "DEFAULT '[]'", "JSON reference URLs"),
        ("cves", "cwe_ids", "TEXT", "DEFAULT '[]'", "JSON CWE list"),
        ("cves", "updated_at", "TEXT", "DEFAULT datetime('now')", "Row update time"),
        ("cves", "cpe_matches", "TEXT", "DEFAULT '[]'", "JSON CPE match objects (migration)"),
        # ioc_cache
        ("ioc_cache", "value", "TEXT", "PRIMARY KEY", "Normalized IOC value"),
        ("ioc_cache", "ioc_type", "TEXT", "NOT NULL", "ip/hash/domain"),
        ("ioc_cache", "result", "TEXT", "NOT NULL", "JSON enrichment result"),
        ("ioc_cache", "cached_at", "TEXT", "DEFAULT datetime('now')", "TTL anchor (6h reads)"),
        # kev_deadlines
        ("kev_deadlines", "cve_id", "TEXT", "PRIMARY KEY", "CVE ID"),
        ("kev_deadlines", "product", "TEXT", "", "KEV product name"),
        ("kev_deadlines", "short_description", "TEXT", "", "CISA short text"),
        ("kev_deadlines", "required_action", "TEXT", "", "Remediation action"),
        ("kev_deadlines", "due_date", "TEXT", "", "Federal due date"),
        ("kev_deadlines", "date_added", "TEXT", "", "KEV catalog add date"),
        ("kev_deadlines", "updated_at", "TEXT", "DEFAULT datetime('now')", "Sync timestamp"),
        # api_usage
        ("api_usage", "service", "TEXT", "NOT NULL, PK part", "Service slug e.g. nvd"),
        ("api_usage", "date_utc", "TEXT", "NOT NULL, PK part", "YYYY-MM-DD"),
        ("api_usage", "month_utc", "TEXT", "NOT NULL", "YYYY-MM"),
        ("api_usage", "count", "INTEGER", "DEFAULT 0", "Calls that day"),
        # sync_state
        ("sync_state", "key", "TEXT", "PRIMARY KEY", "e.g. nvd_last_mod_end"),
        ("sync_state", "value", "TEXT", "NOT NULL", "Watermark value"),
        ("sync_state", "updated_at", "TEXT", "DEFAULT datetime('now')", ""),
        # mitre_techniques
        ("mitre_techniques", "technique_id", "TEXT", "PRIMARY KEY", "e.g. T1190"),
        ("mitre_techniques", "name", "TEXT", "NOT NULL", "Technique name"),
        ("mitre_techniques", "description", "TEXT", "DEFAULT ''", ""),
        ("mitre_techniques", "tactic", "TEXT", "DEFAULT ''", "Tactic name"),
        ("mitre_techniques", "url", "TEXT", "NOT NULL", "attack.mitre.org URL"),
        ("mitre_techniques", "platforms", "TEXT", "DEFAULT '[]'", "JSON platforms"),
        ("mitre_techniques", "detection", "TEXT", "DEFAULT ''", "Detection guidance"),
        # cve_technique_map
        ("cve_technique_map", "cve_id", "TEXT", "NOT NULL, PK part", ""),
        ("cve_technique_map", "technique_id", "TEXT", "NOT NULL, PK part", "FK → mitre_techniques"),
        # atlas_techniques
        ("atlas_techniques", "technique_id", "TEXT", "PRIMARY KEY", "e.g. AML.T0051"),
        ("atlas_techniques", "name", "TEXT", "NOT NULL", ""),
        ("atlas_techniques", "description", "TEXT", "DEFAULT ''", ""),
        ("atlas_techniques", "tactic", "TEXT", "DEFAULT ''", ""),
        ("atlas_techniques", "tactic_id", "TEXT", "DEFAULT ''", ""),
        ("atlas_techniques", "url", "TEXT", "NOT NULL", "atlas.mitre.org URL"),
        # atlas_case_studies
        ("atlas_case_studies", "study_id", "TEXT", "PRIMARY KEY", ""),
        ("atlas_case_studies", "name", "TEXT", "NOT NULL", ""),
        ("atlas_case_studies", "summary", "TEXT", "DEFAULT ''", ""),
        ("atlas_case_studies", "summary_full", "TEXT", "DEFAULT ''", ""),
        ("atlas_case_studies", "techniques", "TEXT", "DEFAULT '[]'", "JSON technique IDs"),
        ("atlas_case_studies", "target", "TEXT", "DEFAULT ''", ""),
        ("atlas_case_studies", "date", "TEXT", "DEFAULT ''", ""),
        ("atlas_case_studies", "study_type", "TEXT", "DEFAULT ''", ""),
        ("atlas_case_studies", "cve_ids", "TEXT", "DEFAULT '[]'", "JSON CVE IDs"),
        # cve_atlas_map
        ("cve_atlas_map", "cve_id", "TEXT", "NOT NULL, PK part", ""),
        ("cve_atlas_map", "technique_id", "TEXT", "NOT NULL, PK part", "FK → atlas_techniques"),
        # epss_history
        ("epss_history", "cve_id", "TEXT", "NOT NULL, PK part", ""),
        ("epss_history", "score", "REAL", "NOT NULL", "EPSS at snapshot"),
        ("epss_history", "recorded_date", "TEXT", "NOT NULL, PK part", ""),
        # cve_exploits
        ("cve_exploits", "id", "INTEGER", "PRIMARY KEY AUTOINCREMENT", ""),
        ("cve_exploits", "cve_id", "TEXT", "NOT NULL", ""),
        ("cve_exploits", "title", "TEXT", "NOT NULL DEFAULT ''", ""),
        ("cve_exploits", "type", "TEXT", "NOT NULL DEFAULT 'poc'", ""),
        ("cve_exploits", "source", "TEXT", "NOT NULL DEFAULT ''", ""),
        ("cve_exploits", "url", "TEXT", "NOT NULL DEFAULT ''", ""),
        ("cve_exploits", "published_date", "TEXT", "DEFAULT ''", ""),
        ("cve_exploits", "fetched_at", "TEXT", "DEFAULT datetime('now')", ""),
        # feed_cache
        ("feed_cache", "cache_key", "TEXT", "PRIMARY KEY", "e.g. sploitus:CVE-..."),
        ("feed_cache", "result", "TEXT", "NOT NULL", "JSON blob"),
        ("feed_cache", "cached_at", "TEXT", "DEFAULT datetime('now')", "TTL checked at read"),
        # cve_change_history
        ("cve_change_history", "id", "INTEGER", "PRIMARY KEY AUTOINCREMENT", ""),
        ("cve_change_history", "cve_id", "TEXT", "NOT NULL", ""),
        ("cve_change_history", "field_name", "TEXT", "NOT NULL", "Tracked field"),
        ("cve_change_history", "old_value", "TEXT", "NOT NULL DEFAULT ''", ""),
        ("cve_change_history", "new_value", "TEXT", "NOT NULL DEFAULT ''", ""),
        ("cve_change_history", "detected_at", "TEXT", "DEFAULT datetime('now')", ""),
        # otx_cve_pulses
        ("otx_cve_pulses", "cve_id", "TEXT", "NOT NULL, PK part", ""),
        ("otx_cve_pulses", "pulse_id", "TEXT", "NOT NULL, PK part", ""),
        ("otx_cve_pulses", "pulse_name", "TEXT", "NOT NULL DEFAULT ''", ""),
        ("otx_cve_pulses", "author", "TEXT", "DEFAULT ''", ""),
        ("otx_cve_pulses", "created_date", "TEXT", "DEFAULT ''", ""),
        ("otx_cve_pulses", "adversary", "TEXT", "DEFAULT ''", ""),
        ("otx_cve_pulses", "malware_families", "TEXT", "DEFAULT '[]'", "JSON"),
        ("otx_cve_pulses", "ioc_count", "INTEGER", "DEFAULT 0", ""),
        ("otx_cve_pulses", "tags", "TEXT", "DEFAULT '[]'", "JSON"),
        ("otx_cve_pulses", "fetched_at", "TEXT", "DEFAULT datetime('now')", ""),
        # otx_pulse_iocs
        ("otx_pulse_iocs", "pulse_id", "TEXT", "NOT NULL, PK part", ""),
        ("otx_pulse_iocs", "ioc_type", "TEXT", "NOT NULL DEFAULT '', PK part", ""),
        ("otx_pulse_iocs", "ioc_value", "TEXT", "NOT NULL, PK part", ""),
        ("otx_pulse_iocs", "description", "TEXT", "DEFAULT ''", ""),
        ("otx_pulse_iocs", "fetched_at", "TEXT", "DEFAULT datetime('now')", ""),
        # correlation_infrastructure
        ("correlation_infrastructure", "cve_id_a", "TEXT", "NOT NULL, PK part", ""),
        ("correlation_infrastructure", "cve_id_b", "TEXT", "NOT NULL, PK part", ""),
        ("correlation_infrastructure", "shared_ip_count", "INTEGER", "DEFAULT 0", ""),
        ("correlation_infrastructure", "confidence", "TEXT", "DEFAULT 'low'", ""),
        ("correlation_infrastructure", "detected_at", "TEXT", "DEFAULT datetime('now')", ""),
        # correlation_actor
        ("correlation_actor", "cve_id", "TEXT", "NOT NULL, PK part", ""),
        ("correlation_actor", "actor_name", "TEXT", "NOT NULL, PK part", ""),
        ("correlation_actor", "actor_sectors", "TEXT", "DEFAULT '[]'", "JSON"),
        ("correlation_actor", "user_sector_match", "INTEGER", "DEFAULT 0", ""),
        ("correlation_actor", "confidence", "TEXT", "DEFAULT 'low'", ""),
        ("correlation_actor", "detected_at", "TEXT", "DEFAULT datetime('now')", ""),
        # correlation_temporal
        ("correlation_temporal", "vendor", "TEXT", "PRIMARY KEY", ""),
        ("correlation_temporal", "current_week_count", "INTEGER", "DEFAULT 0", ""),
        ("correlation_temporal", "average_weekly_count", "REAL", "DEFAULT 0", ""),
        ("correlation_temporal", "anomaly_score", "REAL", "DEFAULT 0", ""),
        ("correlation_temporal", "detected_at", "TEXT", "DEFAULT datetime('now')", ""),
        # mitre_groups
        ("mitre_groups", "group_id", "TEXT", "PRIMARY KEY", "e.g. G0016"),
        ("mitre_groups", "name", "TEXT", "NOT NULL", ""),
        ("mitre_groups", "aliases", "TEXT", "DEFAULT '[]'", "JSON"),
        ("mitre_groups", "description", "TEXT", "DEFAULT ''", ""),
        ("mitre_groups", "sectors", "TEXT", "DEFAULT '[]'", "JSON targeted sectors"),
        ("mitre_groups", "url", "TEXT", "DEFAULT ''", ""),
        # group_technique_map
        ("group_technique_map", "group_id", "TEXT", "NOT NULL, PK part", ""),
        ("group_technique_map", "technique_id", "TEXT", "NOT NULL, PK part", ""),
    ]
    for row in schema:
        ws2.append(row)
    autosize(ws2)

    # Sheet 3: Scheduler Jobs
    ws3 = wb.create_sheet("Scheduler Jobs")
    ws3.append(["Job ID", "Schedule", "Fetches from", "Writes to", "Failure behaviour", "Idempotent"])
    jobs = [
        ("nvd_incremental_sync", "Every NVD_SYNC_INTERVAL_HOURS (default 1h)", "NVD API", "cves, sync_state, feed_cache", "Log; watermark not advanced on fail", "Yes"),
        ("kev_metadata_sync", "Every KEV_SYNC_INTERVAL_MINUTES (default 15m)", "CISA KEV", "kev_deadlines, cves.is_kev", "Log error", "Yes"),
        ("epss_score_sync", "Every EPSS_SYNC_INTERVAL_HOURS (default 6h)", "EPSS CSV/API", "cves, epss_history", "Log error", "Yes"),
        ("weekly_mitre_refresh", "Cron Sun MITRE_REFRESH_HOUR (default 02:00)", "MITRE STIX, ATLAS", "mitre_*, atlas_*, maps", "Log error", "Mostly"),
        ("otx_nightly_correlation", "Cron OTX_CORRELATION (default 02:00 IST)", "OTX API", "otx_*, feed_cache", "Skip if no key", "Yes"),
        ("incident_news_refresh", "Every 4 hours", "6 RSS feeds", "feed_cache", "Per-source errors", "Yes"),
        ("nightly_correlation", "Cron CORRELATION (default 01:00 IST)", "OTX + DB", "correlation_*", "Log; lock skip", "Yes"),
    ]
    for row in jobs:
        ws3.append(row)
    autosize(ws3)

    # Sheet 4: External APIs
    ws4 = wb.create_sheet("External APIs")
    ws4.append(["Service", "Endpoint", "Key env var", "Free tier limit", "Fallback"])
    apis = [
        ("NVD", "services.nvd.nist.gov/rest/json/cves/2.0", "NVD_API_KEY", "50/30s with key", "Retry/backoff; anonymous fallback"),
        ("CISA KEV", "known_exploited_vulnerabilities.json", "", "Unlimited", "[]"),
        ("EPSS", "CSV gzip + api.first.org/data/v1/epss", "", "Unlimited", "{}"),
        ("MITRE STIX", "enterprise-attack.json + CTID CSV", "", "Unlimited", "Job fails"),
        ("ATLAS", "GitHub raw YAML + case-studies API", "ATLAS_YAML_URL", "Unlimited", "Job fails"),
        ("Sploitus", "sploitus.com search API", "", "Unpublished", "None/[]"),
        ("GreyNoise", "api.greynoise.io/v3/community", "GREYNOISE_API_KEY", "50/week", "Unknown classification"),
        ("VirusTotal", "virustotal.com/api/v3", "VIRUSTOTAL_API_KEY", "500/day", "Empty fields"),
        ("AbuseIPDB", "api.abuseipdb.com/api/v2/check", "ABUSEIPDB_API_KEY", "1000/day", "Skipped"),
        ("OTX", "otx.alienvault.com/api/v1", "OTX_API_KEY", "10k/month", "[]"),
        ("OSV", "api.osv.dev/v1/query", "", "Unlimited", "[]"),
        ("CIRCL", "cve.circl.lu/api/cve", "", "Unlimited", "No merge"),
        ("MalwareBazaar", "bazaar.abuse.ch/api", "ABUSECH_AUTH_KEY", "Fair use", "None"),
        ("URLhaus", "urlhaus-api.abuse.ch", "ABUSECH_AUTH_KEY", "Fair use", "None"),
        ("Groq", "api.groq.com/openai/v1/chat/completions", "GROQ_API_KEY", "Console quota", "Anthropic/template"),
        ("Anthropic", "api.anthropic.com/v1/messages", "ANTHROPIC_API_KEY", "Console quota", "Template"),
        ("GitHub", "api.github.com/search/code", "GITHUB_TOKEN", "60/hr without token", "[] rules"),
    ]
    for row in apis:
        ws4.append(row)
    autosize(ws4)

    # Sheet 5: Risk Scoring
    ws5 = wb.create_sheet("Risk Scoring")
    ws5.append(["Component", "Weight", "Source file"])
    for row in [
        ("Asset profile", 0.35, "frontend/src/scoring/riskScore.js"),
        ("KEV status", 0.25, "frontend/src/scoring/riskScore.js"),
        ("EPSS", 0.15, "frontend/src/scoring/riskScore.js"),
        ("Exploit availability", 0.10, "frontend/src/scoring/riskScore.js"),
        ("CVSS", 0.10, "frontend/src/scoring/riskScore.js"),
        ("Momentum", 0.05, "backend/scoring/risk.py calculate_momentum"),
    ]:
        ws5.append(row)
    ws5.append([])
    ws5.append(["Momentum signal", "Description", "Max contribution"])
    signals = [
        ("epss_rising", "EPSS delta over 14 snapshots", "0.50"),
        ("otx_pulse", "OTX pulse fetched_at recency", "0.50"),
        ("kev_recent", "KEV added within 7 days", "0.40"),
        ("rapid_exploitation", "KEV within 30d of publish", "0.30"),
    ]
    for row in signals:
        ws5.append(row)
    autosize(ws5)

    # Sheet 6: Feature Completion
    ws6 = wb.create_sheet("Feature Completion")
    ws6.append(["Feature", "Status", "Notes"])
    features = [
        ("NVD incremental ingest", "Complete", "Watermark + upsert"),
        ("KEV sync", "Complete", "15m default"),
        ("EPSS sync", "Complete", "History snapshots"),
        ("CVE feed", "Complete", "Max 50/page"),
        ("CVE detail ATLAS", "Complete", "has_ai_context + atlas_techniques on GET /api/cves/{id}"),
        ("AI/ML alerts stat", "Complete", "GET /api/stats?frameworks= + feed filter"),
        ("Incidents combined feed", "Complete", "GET /api/case-studies/feed parallel load"),
        ("IOC lookup", "Complete", "6h cache"),
        ("Risk score v1.1b", "Complete", "Client-side"),
        ("Correlation", "Complete", "6h cache"),
        ("Detection tab", "Complete", "Sigma/Elastic"),
        ("PDF + AI summary", "Complete", "Groq/Anthropic/template"),
        ("Backups", "Complete", "6h timer, retention 100, auto-restore"),
        ("POST /api/investigation/summary", "Complete", "Legacy alias to generate_executive_summary"),
        ("Authentication", "Not implemented", "Planned v1.2"),
        ("Repository pattern", "Not implemented", "Planned v1.2"),
    ]
    for row in features:
        ws6.append(row)
    autosize(ws6)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
